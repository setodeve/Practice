import glob
import openpyxl

SHEET = "Sheet1"
files=glob.glob('2*.xlsx')

row_num = 4
lists = {}

for i in files: #file読み込み
  file = openpyxl.load_workbook(i)
  ws = file[SHEET]
  while(1): #各行読み出し
    if(ws.cell(row_num, 2).value is None): break # 指摘が空だったら
    if((ws.cell(row_num, 3).value is None)): # 指摘修正が空だったら
     lists[i] = "未完了"
     break
    row_num += 1
  
print(lists) #指摘が未完了の議事録だけ表示
